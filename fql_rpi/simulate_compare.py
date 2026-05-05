"""
Simulation-Based Controller Comparison
=======================================
Thesis : Edge-Intelligent Aquaculture Aerator Control
         Using Progressive Hybrid FQL-DQN with N3IWF LES
Student: Faril Pirwanhadi (M14128104)

Evaluates RB, FQL, and DQN controllers in the SAME virtual environment
(identical scenarios, identical random seeds) for a fair comparison.

No real Pico or serial connection needed — purely simulation.

Usage:
  uv run python3 simulate_compare.py
  uv run python3 simulate_compare.py --episodes 30 --steps 300 --save results/
"""

import argparse
import math
import os
import random

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

from pond_simulator import PondSimulator, ScenarioType, SimConfig
from fql_agent     import FQLAgent, ACTION_OFF, ACTION_LOW, ACTION_MED, ACTION_HIGH
from dqn_agent     import DQNAgent

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
QTABLE_FILE    = os.path.join(BASE_DIR, "qtable.json")
DQN_MODEL_FILE = os.path.join(BASE_DIR, "dqn_model.pt")

ACTION_NAMES = ["OFF", "LOW", "MED", "HIGH"]
ACTION_COST  = {ACTION_OFF: 0.0, ACTION_LOW: 0.3, ACTION_MED: 0.6, ACTION_HIGH: 1.0}


# ── Rule-based controller (mirrors Pico firmware) ────────────────────────── #

def rule_based_action(pH: float, T: float) -> int:
    if pH < 6.0 or pH > 9.5 or T > 35.0: return ACTION_HIGH
    if pH < 6.5 or pH > 8.5 or T > 30.0: return ACTION_MED
    return ACTION_LOW


# ── NH3 fraction ─────────────────────────────────────────────────────────── #

def nh3_fraction(pH: float, T: float) -> float:
    pka = 0.09018 + 2729.92 / (T + 273.15)
    return 1.0 / (1.0 + 10 ** (pka - pH))


# ── Reward function (mirrors fql_agent.py compute_reward) ────────────────── #

def _classify_state(pH: float, T: float) -> str:
    """Classify a state as SAFE, WARNING, or DANGER."""
    if 6.5 <= pH <= 8.5 and T <= 30.0:
        return "SAFE"
    elif pH < 6.0 or pH > 9.5 or T > 34.0 or T < 18.0:
        return "DANGER"
    return "WARNING"


def compute_reward(pH: float, T: float, action: int,
                   pH_next: float, T_next: float) -> float:
    """
    Outcome-based reward — zone-conditional energy penalty.

    Components:
      1. State quality of next state (+2 SAFE, 0 WARNING, -2 DANGER)
      2. Energy penalty — full in SAFE, tiny in WARNING/DANGER so agents
         aren't punished for using MED/HIGH when needed
      3. NH3 toxicity penalty

    Key insight: energy penalty near-zero in stress zones lets DQN see
    that MED/HIGH lead to SAFE faster (via gamma-discounted future reward)
    without the immediate penalty drowning out the signal.
    """
    if action == 0:  # ACTION_OFF
        return -10.0

    zone_now  = _classify_state(pH, T)
    zone_next = _classify_state(pH_next, T_next)

    # 1. State Quality of next state
    r_state = {"SAFE": 2.0, "WARNING": 0.0, "DANGER": -2.0}[zone_next]

    # 2. Energy penalty — graduated by zone severity
    #    SAFE: full cost (LOW suffices)
    #    WARNING: 40% cost (MED acceptable, HIGH wasteful)
    #    DANGER: 5% cost (emergency, any action justified)
    _cost = {1: 0.0, 2: 0.3, 3: 0.7}
    if zone_now == "SAFE":
        energy = _cost.get(action, 0.0)
    elif zone_now == "WARNING":
        energy = _cost.get(action, 0.0) * 0.40
    else:  # DANGER
        energy = _cost.get(action, 0.0) * 0.05

    # 3. NH3 Toxicity Penalty
    pka = 0.09018 + 2729.92 / (T_next + 273.15)
    nh3_frac = 1.0 / (1.0 + 10 ** (pka - pH_next))
    r_nh3 = nh3_frac * 5.0

    return r_state - energy - r_nh3


# ── FQL pretrain using virtual simulator ──────────────────────────────────── #

def pretrain_fql(fql: FQLAgent, sim: PondSimulator, steps: int = 150_000) -> None:
    """Train FQL from scratch using virtual simulator before evaluation."""
    from pond_simulator import ScenarioType, SimConfig

    # 50% NORMAL so FQL strongly associates LOW with safe conditions.
    # Stress scenarios teach MED (WARNING) and HIGH (DANGER).
    # Previously 20% NORMAL caused 25% HIGH bleed-through into safe states.
    # Balanced scenarios to ensure agents learn DANGER states frequently
    _SCENARIO_ORDER = [
        ScenarioType.NORMAL,
        ScenarioType.ACID_CRASH,
        ScenarioType.ALKALINE,
        ScenarioType.HEAT_STRESS,
        ScenarioType.HIGH_NH3,
        ScenarioType.COLD_STRESS,
        ScenarioType.MULTI_STRESS,
    ]
    EPISODE_LEN = 200

    scen_idx   = 0
    steps_left = 0
    ph, temp   = 7.5, 27.0
    prev_action = None

    random.seed(42)
    np.random.seed(42)
    print(f"  Pretraining FQL for {steps:,} virtual steps (episode={EPISODE_LEN}, no NORMAL)...")
    for i in range(steps):
        if steps_left <= 0:
            scenario = _SCENARIO_ORDER[scen_idx % len(_SCENARIO_ORDER)]
            ph, temp = sim.reset(scenario)
            steps_left = EPISODE_LEN
            scen_idx += 1
            prev_action = None

        action      = fql.select_action(ph, temp)
        ph_next, t_next = sim.step(action)
        reward      = fql.compute_reward(ph, temp, action, ph_next, t_next, prev_action)
        fql.update(ph, temp, action, reward, ph_next, t_next)

        prev_action = action
        ph, temp    = ph_next, t_next
        steps_left -= 1

        if (i + 1) % 10_000 == 0:
            stats = fql.get_stats()
            print(f"    step {i+1:6d} | ε={stats['epsilon']:.3f} | "
                  f"AvgR={stats['avg_reward_100']:+.3f} | "
                  f"converged={stats['converged']}")

    fql.epsilon = 0.0  # greedy for evaluation
    print(f"  Pretrain done. Saving Q-table to {QTABLE_FILE}")
    fql.save_qtable(QTABLE_FILE)


# ── DQN virtual training ─────────────────────────────────────────────────── #

def collect_dqn_buffer(fql: FQLAgent, sim: PondSimulator,
                       n_steps: int = 100_000) -> list:
    """Collect transitions using FQL's learned policy with exploration.

    Using FQL's policy (ε=0.15 exploration) instead of pure random gives DQN
    high-quality demonstrations of WHEN to use MED/HIGH.  Q-learning is
    off-policy, so learning from FQL data is perfectly valid and lets the
    neural network discover improvements via Bellman optimality.
    """
    _BUFFER_SCENARIOS = [
        ScenarioType.NORMAL,
        ScenarioType.ACID_CRASH, ScenarioType.ALKALINE, ScenarioType.HEAT_STRESS,
        ScenarioType.NORMAL,
        ScenarioType.COLD_STRESS, ScenarioType.HIGH_NH3, ScenarioType.MULTI_STRESS,
        ScenarioType.NORMAL,
        ScenarioType.ACID_CRASH, ScenarioType.COLD_STRESS, ScenarioType.HIGH_NH3,
    ]
    _VALID = [ACTION_LOW, ACTION_MED, ACTION_HIGH]
    EXPLORE_EPS = 0.15  # 15% random exploration for coverage

    buffer = []
    scen_idx, steps_left = 0, 0
    ph, temp = 7.5, 27.0

    # Temporarily set FQL epsilon for exploration during collection
    old_eps = fql.epsilon
    fql.epsilon = EXPLORE_EPS

    random.seed(123)
    np.random.seed(123)
    print(f"  Collecting DQN buffer ({n_steps:,} steps, FQL policy ε={EXPLORE_EPS})...")
    for i in range(n_steps):
        if steps_left <= 0:
            sc = _BUFFER_SCENARIOS[scen_idx % len(_BUFFER_SCENARIOS)]
            ph, temp = sim.reset(sc)
            steps_left = 200
            scen_idx += 1

        action = fql.select_action(ph, temp)  # FQL policy with exploration
        ph_next, t_next = sim.step(action)
        r = compute_reward(ph, temp, action, ph_next, t_next)

        buffer.append({"s": [ph, temp], "a": action,
                       "r": round(r, 5), "s_next": [ph_next, t_next]})
        ph, temp = ph_next, t_next
        steps_left -= 1

        if (i + 1) % 10_000 == 0:
            print(f"    buffer {i+1:,}/{n_steps:,}")

    fql.epsilon = old_eps  # restore FQL epsilon
    return buffer


def train_dqn_virtual(fql: FQLAgent, sim: PondSimulator,
                      save_path: str, epochs: int = 20000):
    """Train DQN from virtual buffer using new reward function."""
    buffer = collect_dqn_buffer(fql, sim, n_steps=100_000)
    print(f"  Training DQN ({epochs} epochs on {len(buffer):,} transitions)...")
    try:
        from train_dqn import train_pytorch, train_numpy, TORCH_AVAILABLE
        if TORCH_AVAILABLE:
            train_pytorch(buffer, epochs, save_path)
        else:
            train_numpy(buffer, epochs, save_path)
        dqn = DQNAgent()
        if dqn.load(save_path):
            print("  [DQN] Trained and loaded.")
            return dqn
    except Exception as e:
        print(f"  [DQN] Training failed: {e}")
    return None


# ── Single episode evaluation ─────────────────────────────────────────────── #

def run_episode(controller, sim: PondSimulator, scenario: ScenarioType,
                steps: int, seed: int) -> dict:
    """
    Run one episode and return metrics dict.
    controller: callable (pH, T) -> action int
    """
    random.seed(seed)

    ph, temp = sim.reset(scenario)

    rewards, nh3s, energies, actions_taken = [], [], [], []

    for _ in range(steps):
        action         = controller(ph, temp)
        ph_next, t_next = sim.step(action)

        r = compute_reward(ph, temp, action, ph_next, t_next)

        rewards.append(r)
        nh3s.append(nh3_fraction(ph, temp) * 100.0)
        energies.append(ACTION_COST[action])
        actions_taken.append(action)

        ph, temp = ph_next, t_next

    ph_arr = np.array([sim.ph])   # final pH — use step-by-step for full trace
    # Re-run to collect pH trace (already done above, collect inline next pass)
    # Use rewards/nh3 collected above
    n = len(rewards)
    return {
        "avg_reward":    float(np.mean(rewards)),
        "avg_nh3":       float(np.mean(nh3s)),
        "avg_energy":    float(np.mean(energies)),
        "total_energy":  float(np.sum(energies)),
        "action_dist":   [actions_taken.count(a) / n * 100 for a in range(4)],
        "n":             n,
    }


def run_episode_full(controller, sim: PondSimulator, scenario: ScenarioType,
                     steps: int, seed: int) -> dict:
    """Run episode and return full time series for plotting."""
    random.seed(seed)
    ph, temp = sim.reset(scenario)

    phs, temps, nh3s, rewards, energies, actions_taken = [], [], [], [], [], []

    for _ in range(steps):
        action          = controller(ph, temp)
        ph_next, t_next = sim.step(action)
        r = compute_reward(ph, temp, action, ph_next, t_next)

        phs.append(ph); temps.append(temp)
        nh3s.append(nh3_fraction(ph, temp) * 100.0)
        rewards.append(r)
        energies.append(ACTION_COST[action])
        actions_taken.append(action)

        ph, temp = ph_next, t_next

    n = len(rewards)
    ph_arr = np.array(phs)
    return {
        "pH":           np.array(phs),
        "T":            np.array(temps),
        "NH3":          np.array(nh3s),
        "reward":       np.array(rewards),
        "energy":       np.array(energies),
        "actions":      actions_taken,
        "avg_reward":   float(np.mean(rewards)),
        "avg_nh3":      float(np.mean(nh3s)),
        "avg_energy":   float(np.mean(energies)),
        "ph_safe_pct":  float(((ph_arr >= 6.5) & (ph_arr <= 8.5)).mean() * 100),
        "action_dist":  [actions_taken.count(a) / n * 100 for a in range(4)],
        "n":            n,
    }


# ── Aggregate over multiple episodes ──────────────────────────────────────── #

def evaluate(controller, sim: PondSimulator, scenarios: list,
             episodes: int, steps: int, base_seed: int = 42) -> dict:
    """Run multiple episodes across all scenarios, return aggregated metrics."""
    all_rewards, all_nh3, all_energy, all_actions = [], [], [], []
    all_ph_safe = []

    for ep in range(episodes):
        for scen in scenarios:
            seed = base_seed + ep * 100 + scen.value
            res  = run_episode_full(controller, sim, scen, steps, seed)
            all_rewards.append(res["avg_reward"])
            all_nh3.append(res["avg_nh3"])
            all_energy.append(res["avg_energy"])
            all_ph_safe.append(res["ph_safe_pct"])
            all_actions.extend(res["actions"])

    n_total = len(all_actions)
    return {
        "avg_reward":   float(np.mean(all_rewards)),
        "std_reward":   float(np.std(all_rewards)),
        "avg_nh3":      float(np.mean(all_nh3)),
        "avg_energy":   float(np.mean(all_energy)),
        "ph_safe_pct":  float(np.mean(all_ph_safe)),
        "action_dist":  [all_actions.count(a) / n_total * 100 for a in range(4)],
        "n_episodes":   episodes * len(scenarios),
    }


# ── Time-series for plotting (one representative episode per scenario) ─────── #

def collect_timeseries(controller, sim: PondSimulator, scenario: ScenarioType,
                       steps: int, seed: int = 42) -> dict:
    return run_episode_full(controller, sim, scenario, steps, seed)


# ── Print summary table ───────────────────────────────────────────────────── #

def print_summary(results: dict) -> None:
    controllers = list(results.keys())
    print("\n" + "=" * 65)
    print("  SIMULATION COMPARISON  (same virtual environment)")
    print("=" * 65)

    for name, m in results.items():
        print(f"\n  ── {name} ({m['n_episodes']} episodes) ──")
        print(f"    Avg reward      : {m['avg_reward']:+.4f} ± {m['std_reward']:.4f}")
        print(f"    Avg NH3%%        : {m['avg_nh3']:.3f}%%")
        print(f"    Avg energy/step : {m['avg_energy']:.3f}")
        print(f"    %% time SAFE pH  : {m['ph_safe_pct']:.1f}%%")
        print(f"    Action dist     : " +
              "  ".join(f"{ACTION_NAMES[a]}={m['action_dist'][a]:.1f}%%" for a in range(4)))

    # Pairwise deltas
    names = controllers
    if len(names) >= 2:
        rb  = results.get("Rule-Based")
        fql = results.get("FQL")
        dqn = results.get("DQN")

        def delta(a, b, key):
            return a[key] - b[key]

        if rb and fql:
            print(f"\n  ── FQL vs RB ──")
            print(f"    Reward : {fql['avg_reward']:+.4f} vs {rb['avg_reward']:+.4f}  →  Δ={delta(fql,rb,'avg_reward'):+.4f}")
            print(f"    Energy : {fql['avg_energy']:.3f} vs {rb['avg_energy']:.3f}  →  Δ={delta(fql,rb,'avg_energy'):+.3f}")
            print(f"    NH3%%   : {fql['avg_nh3']:.3f} vs {rb['avg_nh3']:.3f}  →  Δ={delta(fql,rb,'avg_nh3'):+.3f}")

        if rb and dqn:
            print(f"\n  ── DQN vs RB ──")
            print(f"    Reward : {dqn['avg_reward']:+.4f} vs {rb['avg_reward']:+.4f}  →  Δ={delta(dqn,rb,'avg_reward'):+.4f}")
            print(f"    Energy : {dqn['avg_energy']:.3f} vs {rb['avg_energy']:.3f}  →  Δ={delta(dqn,rb,'avg_energy'):+.3f}")
            print(f"    NH3%%   : {dqn['avg_nh3']:.3f} vs {rb['avg_nh3']:.3f}  →  Δ={delta(dqn,rb,'avg_nh3'):+.3f}")

        if fql and dqn:
            print(f"\n  ── DQN vs FQL ──")
            print(f"    Reward : {dqn['avg_reward']:+.4f} vs {fql['avg_reward']:+.4f}  →  Δ={delta(dqn,fql,'avg_reward'):+.4f}")
            print(f"    Energy : {dqn['avg_energy']:.3f} vs {fql['avg_energy']:.3f}  →  Δ={delta(dqn,fql,'avg_energy'):+.3f}")
            print(f"    NH3%%   : {dqn['avg_nh3']:.3f} vs {fql['avg_nh3']:.3f}  →  Δ={delta(dqn,fql,'avg_nh3'):+.3f}")

    print("=" * 65 + "\n")


# ── Plots ─────────────────────────────────────────────────────────────────── #

def rolling(arr, w=20):
    return np.convolve(arr, np.ones(w) / w, mode="same")


def plot_comparison(ts_results: dict, save_dir: str | None = None) -> None:
    """
    ts_results: {controller_name: {scenario_label: timeseries_dict}}
    """
    colors = {"Rule-Based": "#e74c3c", "FQL": "#27ae60", "DQN": "#2980b9"}
    scenarios = list(next(iter(ts_results.values())).keys())

    n_scen = len(scenarios)
    fig = plt.figure(figsize=(18, 4 * n_scen))
    fig.suptitle("Simulation Comparison: RB vs FQL vs DQN\n(identical virtual environment)",
                 fontsize=13, fontweight="bold")

    gs = gridspec.GridSpec(n_scen, 3, figure=fig, hspace=0.5, wspace=0.35)

    for row, scen_label in enumerate(scenarios):
        # pH
        ax = fig.add_subplot(gs[row, 0])
        for name, scen_data in ts_results.items():
            d = scen_data[scen_label]
            ax.plot(d["pH"], color=colors[name], linewidth=0.8, alpha=0.85, label=name)
        ax.axhline(6.5, color="red", linestyle=":", linewidth=0.7)
        ax.axhline(8.5, color="red", linestyle=":", linewidth=0.7)
        ax.fill_between(range(len(d["pH"])), 6.5, 8.5, alpha=0.05, color="green")
        ax.set_title(f"{scen_label} — pH"); ax.set_ylabel("pH"); ax.grid(True, alpha=0.3)
        if row == 0: ax.legend(fontsize=7)

        # NH3
        ax = fig.add_subplot(gs[row, 1])
        for name, scen_data in ts_results.items():
            d = scen_data[scen_label]
            ax.plot(d["NH3"], color=colors[name], linewidth=0.8, alpha=0.85, label=name)
        ax.set_title(f"{scen_label} — NH3 (%)"); ax.set_ylabel("NH3 %"); ax.grid(True, alpha=0.3)

        # Cumulative reward
        ax = fig.add_subplot(gs[row, 2])
        for name, scen_data in ts_results.items():
            d = scen_data[scen_label]
            ax.plot(np.cumsum(d["reward"]), color=colors[name], linewidth=0.8,
                    alpha=0.85, label=name)
        ax.set_title(f"{scen_label} — Cumulative Reward")
        ax.set_ylabel("Cum. Reward"); ax.grid(True, alpha=0.3)

    plt.tight_layout(rect=[0, 0, 1, 0.96])

    if save_dir:
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir, "simulation_comparison.png")
        plt.savefig(path, dpi=150, bbox_inches="tight")
        print(f"Plot saved: {path}")
    plt.show()


def plot_bar_summary(results: dict, save_dir: str | None = None) -> None:
    """Bar chart summary: reward, energy, NH3, pH safe across all controllers."""
    names   = list(results.keys())
    colors  = ["#e74c3c", "#27ae60", "#2980b9"]
    metrics = {
        "Avg Reward":      [results[n]["avg_reward"]   for n in names],
        "Avg Energy/step": [results[n]["avg_energy"]   for n in names],
        "Avg NH3 %":       [results[n]["avg_nh3"]      for n in names],
        "pH Safe %":       [results[n]["ph_safe_pct"]  for n in names],
    }

    fig, axes = plt.subplots(1, 4, figsize=(16, 5))
    fig.suptitle("Simulation Summary: RB vs FQL vs DQN",
                 fontsize=13, fontweight="bold")

    for ax, (title, vals) in zip(axes, metrics.items()):
        bars = ax.bar(names, vals, color=colors[:len(names)], alpha=0.85, edgecolor="white")
        ax.set_title(title); ax.set_ylabel(title); ax.grid(True, alpha=0.3, axis="y")
        for bar, val in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + abs(bar.get_height()) * 0.02,
                    f"{val:.3f}", ha="center", va="bottom", fontsize=9)

    # Action distribution
    fig2, ax2 = plt.subplots(figsize=(10, 5))
    x = np.arange(4)
    w = 0.25
    for i, name in enumerate(names):
        offset = (i - len(names) / 2 + 0.5) * w
        ax2.bar(x + offset, results[name]["action_dist"], w,
                label=name, color=colors[i], alpha=0.85)
    ax2.set_xticks(x); ax2.set_xticklabels(ACTION_NAMES)
    ax2.set_ylabel("Usage (%)"); ax2.set_title("Action Distribution")
    ax2.legend(); ax2.grid(True, alpha=0.3, axis="y")
    fig2.suptitle("Action Distribution: RB vs FQL vs DQN", fontweight="bold")

    plt.tight_layout()

    if save_dir:
        os.makedirs(save_dir, exist_ok=True)
        p1 = os.path.join(save_dir, "sim_metrics_bar.png")
        p2 = os.path.join(save_dir, "sim_action_dist.png")
        fig.savefig(p1, dpi=150, bbox_inches="tight")
        fig2.savefig(p2, dpi=150, bbox_inches="tight")
        print(f"Plots saved: {p1}, {p2}")
    plt.show()


# ── Per-scenario evaluation ───────────────────────────────────────────────── #

def evaluate_per_scenario(controllers, sim, scenarios, episodes=20,
                          steps=300, seed=42):
    """Returns {ctrl_name: {scenario_label: metrics_dict}}"""
    results = {}
    for name, ctrl in controllers.items():
        results[name] = {}
        for scen in scenarios:
            rews, nh3s, engs, safes, acts = [], [], [], [], []
            for ep in range(episodes):
                s = seed + ep * 100 + scen.value
                res = run_episode_full(ctrl, sim, scen, steps, s)
                rews.append(res["avg_reward"])
                nh3s.append(res["avg_nh3"])
                engs.append(res["avg_energy"])
                safes.append(res["ph_safe_pct"])
                acts.extend(res["actions"])
            n = len(acts)
            results[name][PondSimulator.label(scen)] = {
                "avg_reward": np.mean(rews), "std_reward": np.std(rews),
                "rewards": rews,
                "avg_nh3": np.mean(nh3s), "avg_energy": np.mean(engs),
                "ph_safe_pct": np.mean(safes),
                "action_dist": [acts.count(a)/n*100 for a in range(4)],
            }
    return results


# ── CSV & LaTeX exports ───────────────────────────────────────────────────── #

def export_csv(per_scen, path):
    import csv
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Scenario","Controller","Avg Reward","Std Reward",
                     "Avg NH3 %","Avg Energy","pH Safe %",
                     "OFF %","LOW %","MED %","HIGH %"])
        for ctrl in per_scen:
            for scen, m in per_scen[ctrl].items():
                w.writerow([scen, ctrl,
                    f"{m['avg_reward']:.4f}", f"{m['std_reward']:.4f}",
                    f"{m['avg_nh3']:.3f}", f"{m['avg_energy']:.3f}",
                    f"{m['ph_safe_pct']:.1f}",
                    *[f"{m['action_dist'][a]:.1f}" for a in range(4)]])
    print(f"  CSV saved: {path}")


def export_excel(agg, per_scen, path):
    """Export results to Excel with Summary and Per-Scenario sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  ⚠ openpyxl not installed, skipping Excel export.")
        print("    Install with: pip install openpyxl")
        return

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_w = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    # Title
    ws1["A1"] = "Simulation Comparison Results"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:G1")

    # Headers
    headers = ["Controller", "Avg Reward", "Std Reward", "Avg Energy/step",
               "NH3 (%)", "pH Safe (%)", "Best Action"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = header_font_w
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for i, (name, m) in enumerate(agg.items(), 4):
        ad = m["action_dist"]
        best = ACTION_NAMES[ad.index(max(ad))]
        vals = [name, round(m["avg_reward"], 4), round(m["std_reward"], 4),
                round(m["avg_energy"], 3), round(m["avg_nh3"], 3),
                round(m["ph_safe_pct"], 1), best]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Delta row
    ctrls = list(agg.keys())
    if "DQN" in agg and "Rule-Based" in agg:
        r = len(agg) + 5
        ws1.cell(row=r, column=1, value="DQN vs RB (Δ)").font = Font(bold=True, italic=True)
        d, rb = agg["DQN"], agg["Rule-Based"]
        ws1.cell(row=r, column=2, value=round(d["avg_reward"]-rb["avg_reward"], 4))
        ws1.cell(row=r, column=4, value=round(d["avg_energy"]-rb["avg_energy"], 3))
        ws1.cell(row=r, column=5, value=round(d["avg_nh3"]-rb["avg_nh3"], 3))
        ws1.cell(row=r, column=6, value=round(d["ph_safe_pct"]-rb["ph_safe_pct"], 1))

    # Action distribution sub-table
    r = len(agg) + 7
    ws1.cell(row=r, column=1, value="Action Distribution (%)").font = Font(bold=True, size=12)
    for col, h in enumerate(["Controller","OFF","LOW","MED","HIGH"], 1):
        cell = ws1.cell(row=r+1, column=col, value=h)
        cell.font = header_font_w; cell.fill = header_fill; cell.border = thin_border
    for i, (name, m) in enumerate(agg.items(), r+2):
        ws1.cell(row=i, column=1, value=name).border = thin_border
        for j, v in enumerate(m["action_dist"]):
            cell = ws1.cell(row=i, column=j+2, value=round(v, 1))
            cell.border = thin_border; cell.alignment = Alignment(horizontal="center")

    # Column widths
    for col in range(1, 8):
        ws1.column_dimensions[chr(64+col)].width = 18

    # ── Sheet 2: Per-Scenario ──
    ws2 = wb.create_sheet("Per-Scenario")
    headers2 = ["Scenario","Controller","Avg Reward","Std Reward",
                "NH3 (%)","Energy/step","pH Safe (%)",
                "OFF %","LOW %","MED %","HIGH %"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_w; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = thin_border

    row = 2
    for ctrl in per_scen:
        for scen, m in per_scen[ctrl].items():
            vals = [scen, ctrl, round(m["avg_reward"],4), round(m["std_reward"],4),
                    round(m["avg_nh3"],3), round(m["avg_energy"],3),
                    round(m["ph_safe_pct"],1),
                    *[round(m["action_dist"][a],1) for a in range(4)]]
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

    for col in range(1, 12):
        ws2.column_dimensions[chr(64+col) if col < 27 else "A" + chr(64+col-26)].width = 16

    wb.save(path)
    print(f"  Excel saved: {path}")


def export_latex_table(agg, path):
    with open(path, "w") as f:
        f.write("\\begin{table}[h]\n\\centering\n")
        f.write("\\caption{Simulation Comparison Results}\n\\label{tab:sim}\n")
        f.write("\\begin{tabular}{lcccc}\n\\hline\n")
        f.write("Controller & Avg Reward & Energy/step & NH3 (\\%) & pH Safe (\\%) \\\\\n\\hline\n")
        for name, m in agg.items():
            f.write(f"{name} & {m['avg_reward']:+.4f} $\\pm$ {m['std_reward']:.4f} "
                    f"& {m['avg_energy']:.3f} & {m['avg_nh3']:.3f} "
                    f"& {m['ph_safe_pct']:.1f} \\\\\n")
        f.write("\\hline\n\\end{tabular}\n\\end{table}\n")
    print(f"  LaTeX saved: {path}")


# ── Boxplot ───────────────────────────────────────────────────────────────── #

def plot_boxplot(per_scen, save_dir):
    _C = {"Rule-Based": "#e74c3c", "FQL": "#27ae60", "DQN": "#2980b9"}
    ctrls = list(per_scen.keys())
    scens = list(list(per_scen.values())[0].keys())
    fig, axes = plt.subplots(1, len(scens), figsize=(3.5*len(scens), 5), sharey=True)
    fig.suptitle("Reward Distribution per Scenario", fontsize=14, fontweight="bold")
    for i, scen in enumerate(scens):
        ax = axes[i]
        data = [per_scen[c][scen]["rewards"] for c in ctrls]
        bp = ax.boxplot(data, labels=ctrls, patch_artist=True, widths=0.6)
        for patch, c in zip(bp["boxes"], ctrls):
            patch.set_facecolor(_C[c]); patch.set_alpha(0.7)
        ax.set_title(scen, fontsize=8); ax.grid(True, alpha=0.3, axis="y")
        ax.axhline(0, color="gray", ls="--", lw=0.5)
        if i == 0: ax.set_ylabel("Avg Reward")
    plt.tight_layout()
    p = os.path.join(save_dir, "boxplot_reward.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ── Per-scenario grouped bars ─────────────────────────────────────────────── #

def plot_per_scenario_bars(per_scen, save_dir):
    _C = {"Rule-Based": "#e74c3c", "FQL": "#27ae60", "DQN": "#2980b9"}
    ctrls = list(per_scen.keys())
    scens = list(list(per_scen.values())[0].keys())
    metrics = [("Avg Reward","avg_reward"),("pH Safe %","ph_safe_pct"),
               ("Avg Energy","avg_energy"),("NH3 %","avg_nh3")]
    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle("Per-Scenario Performance", fontsize=14, fontweight="bold")
    for ax, (title, key) in zip(axes.flat, metrics):
        x = np.arange(len(scens)); w = 0.25
        for j, c in enumerate(ctrls):
            vals = [per_scen[c][s][key] for s in scens]
            ax.bar(x + (j - len(ctrls)/2 + 0.5)*w, vals, w, label=c, color=_C[c], alpha=0.85)
        ax.set_xticks(x); ax.set_xticklabels(scens, rotation=30, ha="right", fontsize=8)
        ax.set_title(title); ax.legend(fontsize=8); ax.grid(True, alpha=0.3, axis="y")
    plt.tight_layout()
    p = os.path.join(save_dir, "per_scenario_bars.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ── Radar chart ───────────────────────────────────────────────────────────── #

def plot_radar(agg, save_dir):
    _C = {"Rule-Based": "#e74c3c", "FQL": "#27ae60", "DQN": "#2980b9"}
    ctrls = list(agg.keys())
    labels = ["Reward","Energy\nEfficiency","NH3\nReduction","pH Safety"]
    raw = {}
    for c in ctrls:
        m = agg[c]
        raw[c] = [m["avg_reward"], 1.0-m["avg_energy"],
                  1.0-m["avg_nh3"]/100, m["ph_safe_pct"]/100]
    n = len(labels)
    mins = [min(raw[c][i] for c in ctrls) for i in range(n)]
    maxs = [max(raw[c][i] for c in ctrls) for i in range(n)]
    scaled = {}
    for c in ctrls:
        scaled[c] = [(raw[c][i]-mins[i])/(maxs[i]-mins[i])
                     if maxs[i]-mins[i] > 0 else 0.5 for i in range(n)]
    angles = np.linspace(0, 2*np.pi, n, endpoint=False).tolist() + [0]
    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
    ax.set_title("Multi-Metric Radar", fontsize=13, fontweight="bold", pad=20)
    for c in ctrls:
        vals = scaled[c] + scaled[c][:1]
        ax.plot(angles, vals, "o-", lw=2, label=c, color=_C[c])
        ax.fill(angles, vals, alpha=0.15, color=_C[c])
    ax.set_xticks(angles[:-1]); ax.set_xticklabels(labels, fontsize=10)
    ax.set_ylim(0, 1.1); ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))
    p = os.path.join(save_dir, "radar_comparison.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ── Zone distribution ─────────────────────────────────────────────────────── #

def _classify_zone(ph, t):
    if 6.5 <= ph <= 8.5 and t <= 30.0: return "SAFE"
    elif ph < 6.0 or ph > 9.5 or t > 34.0 or t < 18.0: return "DANGER"
    return "WARNING"

def plot_zone_dist(controllers, sim, scenarios, save_dir,
                   episodes=20, steps=300, seed=42):
    zone_data = {}
    for name, ctrl in controllers.items():
        zones = {"SAFE": 0, "WARNING": 0, "DANGER": 0}; total = 0
        for ep in range(episodes):
            for scen in scenarios:
                res = run_episode_full(ctrl, sim, scen, steps,
                                       seed + ep*100 + scen.value)
                for ph, t in zip(res["pH"], res["T"]):
                    zones[_classify_zone(ph, t)] += 1; total += 1
        zone_data[name] = {k: v/total*100 for k, v in zones.items()}

    ctrls = list(zone_data.keys())
    zcolors = {"SAFE": "#2ecc71", "WARNING": "#f39c12", "DANGER": "#e74c3c"}
    fig, ax = plt.subplots(figsize=(8, 5))
    fig.suptitle("Time in Each Zone (%)", fontsize=13, fontweight="bold")
    x = np.arange(len(ctrls)); bottom = np.zeros(len(ctrls))
    for z in ["SAFE", "WARNING", "DANGER"]:
        vals = [zone_data[c][z] for c in ctrls]
        ax.bar(x, vals, bottom=bottom, label=z, color=zcolors[z], alpha=0.85)
        for i, v in enumerate(vals):
            if v > 3:
                ax.text(i, bottom[i]+v/2, f"{v:.1f}%", ha="center",
                        va="center", fontsize=10, fontweight="bold")
        bottom += vals
    ax.set_xticks(x); ax.set_xticklabels(ctrls); ax.set_ylabel("%"); ax.legend()
    p = os.path.join(save_dir, "zone_distribution.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ── Policy map heatmap ────────────────────────────────────────────────────── #

def plot_policy_maps(controllers, save_dir):
    from matplotlib.colors import ListedColormap, BoundaryNorm
    from matplotlib.patches import Patch
    ph_r = np.linspace(5.0, 10.0, 50); t_r = np.linspace(16.0, 36.0, 40)
    cmap = ListedColormap(["#3498db", "#f39c12", "#e74c3c"])
    norm = BoundaryNorm([0.5, 1.5, 2.5, 3.5], cmap.N)
    ctrls = list(controllers.keys())
    fig, axes = plt.subplots(1, len(ctrls), figsize=(6*len(ctrls), 5))
    fig.suptitle("Policy Map — Action at (pH, T)", fontsize=14, fontweight="bold")
    if len(ctrls) == 1: axes = [axes]
    for ax, name in zip(axes, ctrls):
        grid = np.zeros((len(t_r), len(ph_r)))
        for i, t in enumerate(t_r):
            for j, ph in enumerate(ph_r):
                grid[i, j] = controllers[name](ph, t)
        ax.imshow(grid, aspect="auto", origin="lower",
                  extent=[5.0, 10.0, 16.0, 36.0], cmap=cmap, norm=norm,
                  interpolation="nearest")
        ax.axvline(6.5, color="white", ls="--", lw=1.2, alpha=0.8)
        ax.axvline(8.5, color="white", ls="--", lw=1.2, alpha=0.8)
        ax.axhline(30.0, color="white", ls="--", lw=1.2, alpha=0.8)
        ax.set_xlabel("pH"); ax.set_ylabel("Temp (°C)"); ax.set_title(name)
    legend_el = [Patch(facecolor="#3498db", label="LOW"),
                 Patch(facecolor="#f39c12", label="MED"),
                 Patch(facecolor="#e74c3c", label="HIGH")]
    fig.legend(handles=legend_el, loc="lower center", ncol=3, fontsize=10)
    plt.tight_layout(rect=[0, 0.05, 1, 0.95])
    p = os.path.join(save_dir, "policy_map_heatmap.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ── Action timeline ───────────────────────────────────────────────────────── #

def plot_action_timeline(controllers, sim, save_dir, steps=300, seed=42):
    key_scens = [ScenarioType.NORMAL, ScenarioType.ACID_CRASH,
                 ScenarioType.ALKALINE, ScenarioType.HIGH_NH3]
    act_c = {0:"#95a5a6", 1:"#3498db", 2:"#f39c12", 3:"#e74c3c"}
    ctrls = list(controllers.keys())
    fig, axes = plt.subplots(len(key_scens), len(ctrls),
                             figsize=(5*len(ctrls), 2.5*len(key_scens)), sharex=True)
    fig.suptitle("Action Timeline per Step", fontsize=13, fontweight="bold")
    for row, scen in enumerate(key_scens):
        for col, name in enumerate(ctrls):
            ax = axes[row][col]
            res = run_episode_full(controllers[name], sim, scen, steps, seed)
            for i, a in enumerate(res["actions"]):
                ax.bar(i, 1, color=act_c[a], width=1.0)
            ax.set_yticks([])
            if row == 0: ax.set_title(name, fontsize=10)
            if col == 0: ax.set_ylabel(PondSimulator.label(scen), fontsize=8)
            ax.set_xlim(0, steps)
    from matplotlib.patches import Patch
    leg = [Patch(facecolor=act_c[a], label=ACTION_NAMES[a]) for a in [1,2,3]]
    fig.legend(handles=leg, loc="lower center", ncol=3, fontsize=9)
    plt.tight_layout(rect=[0, 0.04, 1, 0.96])
    p = os.path.join(save_dir, "action_timeline.png")
    fig.savefig(p, dpi=150, bbox_inches="tight"); plt.close(fig)
    print(f"  Plot: {p}")


# ══════════════════════════════════════════════════════════════════════════ #
#  Entry point
# ══════════════════════════════════════════════════════════════════════════ #

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Simulate RB vs FQL vs DQN comparison")
    parser.add_argument("--episodes", type=int, default=20)
    parser.add_argument("--steps",    type=int, default=300)
    parser.add_argument("--seed",     type=int, default=42)
    parser.add_argument("--scenarios", nargs="+",
                        choices=["all","normal","acid","alkaline","cold","heat","nh3","multi"],
                        default=["all"])
    parser.add_argument("--pretrain-steps", type=int, default=80_000)
    parser.add_argument("--retrain-dqn", action="store_true")
    args = parser.parse_args()

    _SCEN_MAP = {
        "normal": ScenarioType.NORMAL, "acid": ScenarioType.ACID_CRASH,
        "alkaline": ScenarioType.ALKALINE, "cold": ScenarioType.COLD_STRESS,
        "heat": ScenarioType.HEAT_STRESS, "nh3": ScenarioType.HIGH_NH3,
        "multi": ScenarioType.MULTI_STRESS,
    }
    scenarios = list(ScenarioType) if "all" in args.scenarios else \
                [_SCEN_MAP[s] for s in args.scenarios]

    SAVE_DIR = os.path.join(BASE_DIR, "results", "simulation")
    os.makedirs(SAVE_DIR, exist_ok=True)

    print("=" * 65)
    print("  Simulation-Based Controller Evaluation")
    print(f"  Episodes : {args.episodes} per scenario × {len(scenarios)} scenarios")
    print(f"  Steps    : {args.steps} per episode")
    print(f"  Scenarios: {[PondSimulator.label(s) for s in scenarios]}")
    print("=" * 65)

    # ── Load controllers ──
    sim = PondSimulator(SimConfig())
    rb_controller = rule_based_action
    print("  [RB]  Rule-Based controller loaded.")

    fql = FQLAgent()
    if os.path.exists(QTABLE_FILE) and fql.load_qtable(QTABLE_FILE):
        fql.epsilon = 0.0
        print(f"  [FQL] Q-table loaded: {QTABLE_FILE}  (ε=0, greedy)")
    else:
        print("  [FQL] No Q-table found — pretraining from virtual simulator...")
        pretrain_fql(fql, sim, steps=args.pretrain_steps)
    fql_controller = lambda ph, t: fql.select_action(ph, t)

    DQN_VIRTUAL_FILE = os.path.join(BASE_DIR, "dqn_model_virtual.pt")
    dqn = None
    if not args.retrain_dqn and os.path.exists(DQN_VIRTUAL_FILE):
        dqn = DQNAgent()
        if dqn.load(DQN_VIRTUAL_FILE):
            print(f"  [DQN] Virtual model loaded: {DQN_VIRTUAL_FILE}")
        else:
            dqn = None
    if dqn is None:
        print("  [DQN] Training from virtual simulator (new reward function)...")
        dqn = train_dqn_virtual(fql, sim, DQN_VIRTUAL_FILE)
    dqn_controller = (lambda ph, t: dqn.select_action(ph, t)) if dqn else None
    if dqn_controller is None:
        print("  [DQN] Skipped.")

    controllers = {"Rule-Based": rb_controller, "FQL": fql_controller}
    if dqn_controller:
        controllers["DQN"] = dqn_controller
    print()

    # ── Aggregate evaluation ──
    results = {}
    for name, ctrl in controllers.items():
        print(f"  Evaluating {name}...")
        results[name] = evaluate(ctrl, sim, scenarios,
                                 args.episodes, args.steps, args.seed)
    print()
    print_summary(results)

    # ── Per-scenario evaluation ──
    print("  Running per-scenario evaluation...")
    per_scen = evaluate_per_scenario(controllers, sim, scenarios,
                                     args.episodes, args.steps, args.seed)

    # ── Time-series ──
    ts_results = {name: {} for name in controllers}
    for scen in scenarios:
        label = PondSimulator.label(scen)
        for name, ctrl in controllers.items():
            ts_results[name][label] = collect_timeseries(
                ctrl, sim, scen, args.steps, seed=args.seed)

    # ── Generate ALL plots & tables ──
    print(f"\n  Saving all outputs to {SAVE_DIR}/\n")
    plot_comparison(ts_results, save_dir=SAVE_DIR)
    plot_bar_summary(results, save_dir=SAVE_DIR)
    export_csv(per_scen, os.path.join(SAVE_DIR, "per_scenario_results.csv"))
    export_excel(results, per_scen, os.path.join(SAVE_DIR, "simulation_results.xlsx"))
    export_latex_table(results, os.path.join(SAVE_DIR, "latex_table.tex"))
    plot_boxplot(per_scen, SAVE_DIR)
    plot_per_scenario_bars(per_scen, SAVE_DIR)
    plot_radar(results, SAVE_DIR)
    print("  Computing zone distributions...")
    plot_zone_dist(controllers, sim, scenarios, SAVE_DIR,
                   args.episodes, args.steps, args.seed)
    plot_policy_maps(controllers, SAVE_DIR)
    plot_action_timeline(controllers, sim, SAVE_DIR, args.steps, args.seed)

    print(f"\n  ✅ All outputs saved to: {SAVE_DIR}/")
    print("=" * 65)
